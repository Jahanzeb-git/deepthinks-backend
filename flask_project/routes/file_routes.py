import os
import io
import logging
import base64
from flask import Blueprint, request, jsonify, current_app
from auth import token_required

file_bp = Blueprint('file_bp', __name__)

# Graceful imports with fallbacks
try:
    import magic
    HAS_MAGIC = True
except ImportError:
    HAS_MAGIC = False
    logging.warning("python-magic not available. File type detection will be limited.")

try:
    import pypdf
    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False
    logging.warning("pypdf not available. PDF processing will be disabled.")

try:
    import docx
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    logging.warning("python-docx not available. DOCX processing will be disabled.")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logging.warning("openpyxl not available. XLSX processing will be disabled.")

# A dictionary to map MIME types to file extensions
MIME_TYPE_MAP = {
    'text/plain': '.txt',
    'application/pdf': '.pdf',
    'application/msword': '.doc',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
    'text/markdown': '.md',
    'text/csv': '.csv',
    'application/json': '.json',
    'application/xml': '.xml',
    'text/html': '.html',
    'text/css': '.css',
    'application/javascript': '.js',
    'application/x-python-code': '.py',
    'text/x-python': '.py',
    'text/x-c': '.c',
    'text/x-c++src': '.cpp',
    'text/x-java-source': '.java',
    'application/x-sh': '.sh',
    'image/jpeg': '.jpeg',
    'image/png': '.png',
}

# File extension to MIME type mapping (fallback)
EXT_TO_MIME = {
    '.txt': 'text/plain',
    '.pdf': 'application/pdf',
    '.doc': 'application/msword',
    '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    '.md': 'text/markdown',
    '.csv': 'text/csv',
    '.json': 'application/json',
    '.xml': 'application/xml',
    '.html': '.html',
    '.css': '.css',
    '.js': 'application/javascript',
    '.py': 'text/x-python',
    '.c': 'text/x-c',
    '.cpp': 'text/x-c++src',
    '.java': 'text/x-java-source',
    '.sh': 'application/x-sh',
    '.jpg': 'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.png': 'image/png',
}

def get_file_extension(mime_type):
    """Get the file extension from the MIME type."""
    return MIME_TYPE_MAP.get(mime_type)

def detect_mime_type(file_content_bytes, filename):
    """Detect MIME type with fallback methods."""
    # Method 1: Use python-magic if available
    if HAS_MAGIC:
        try:
            return magic.from_buffer(file_content_bytes, mime=True)
        except Exception as e:
            logging.warning(f"Magic detection failed: {e}")

    # Method 2: Fallback to file extension
    if filename:
        _, ext = os.path.splitext(filename.lower())
        if ext in EXT_TO_MIME:
            return EXT_TO_MIME[ext]

    # Method 3: Basic content analysis
    try:
        # Try to decode as text
        file_content_bytes.decode('utf-8')
        return 'text/plain'
    except UnicodeDecodeError:
        pass

    # Check for PDF signature
    if file_content_bytes.startswith(b'%PDF'):
        return 'application/pdf'

    # Check for ZIP-based formats (docx, xlsx)
    if file_content_bytes.startswith(b'PK'):
        if filename:
            if filename.lower().endswith('.docx'):
                return 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            elif filename.lower().endswith('.xlsx'):
                return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Default fallback
    return 'application/octet-stream'

def extract_text_from_pdf(file_content_bytes):
    """Extract text from PDF with error handling."""
    if not HAS_PYPDF:
        return "Error: PDF processing not available. Please install pypdf."

    try:
        pdf_reader = pypdf.PdfReader(io.BytesIO(file_content_bytes))
        text = ""
        for page_num, page in enumerate(pdf_reader.pages):
            try:
                text += page.extract_text()
            except Exception as e:
                logging.warning(f"Failed to extract text from page {page_num}: {e}")
        return text if text.strip() else "Error: Could not extract text from PDF."
    except Exception as e:
        return f"Error extracting text from PDF: {e}"

def extract_text_from_docx(file_content_bytes):
    """Extract text from DOCX with error handling."""
    if not HAS_DOCX:
        return "Error: DOCX processing not available. Please install python-docx."

    try:
        doc = docx.Document(io.BytesIO(file_content_bytes))
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text if text.strip() else "Error: Could not extract text from DOCX."
    except Exception as e:
        return f"Error extracting text from DOCX: {e}"

def extract_text_from_xlsx(file_content_bytes):
    """Extract text from XLSX with error handling."""
    if not HAS_OPENPYXL:
        return "Error: XLSX processing not available. Please install openpyxl."

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content_bytes))
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = ",".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():  # Only add non-empty rows
                    text += row_text + "\n"
            text += "\n"
        return text if text.strip() else "Error: Could not extract text from XLSX."
    except Exception as e:
        return f"Error extracting text from XLSX: {e}"

def extract_text_from_file(file_content_bytes, mime_type):
    """Extract text from file content based on MIME type."""
    try:
        if mime_type == 'application/pdf':
            return extract_text_from_pdf(file_content_bytes)
        elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
            return extract_text_from_docx(file_content_bytes)
        elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            return extract_text_from_xlsx(file_content_bytes)
        else:
            # Try to decode as text
            try:
                return file_content_bytes.decode('utf-8')
            except UnicodeDecodeError:
                # Try with different encodings
                for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                    try:
                        return file_content_bytes.decode(encoding)
                    except UnicodeDecodeError:
                        continue
                return "Error: Failed to decode file content. The file may be binary or in an unsupported encoding."
    except Exception as e:
        return f"Error processing file: {e}"

@file_bp.route('/upload', methods=['POST'])
@token_required
def upload_file(current_user):
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file part"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400

        session_id = request.form.get('session_id')
        if not session_id:
            return jsonify({"error": "session_id is required"}), 400

        # Check file size (limit to 10MB)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning

        if file_size > 10 * 1024 * 1024:  # 10MB limit
            return jsonify({"error": "File size exceeds 10MB limit"}), 400

        if file:
            file_content_bytes = file.read()
            if not file_content_bytes:
                return jsonify({"error": "File is empty"}), 400

            # Detect MIME type with fallbacks
            mime_type = detect_mime_type(file_content_bytes, file.filename)

            user_id = current_user['id']
            cache_key = f"{user_id}-{session_id}"

            file_extension = get_file_extension(mime_type)
            original_filename = os.path.splitext(file.filename)[0]
            filename_with_extension = original_filename + (file_extension or os.path.splitext(file.filename)[1])

            # Ensure file_cache exists
            if not hasattr(current_app, 'file_cache'):
                current_app.file_cache = {}

            if mime_type in ['image/jpeg', 'image/png']:
                encoded_string = base64.b64encode(file_content_bytes).decode('utf-8')
                data_url = f"data:{mime_type};base64,{encoded_string}"

                current_app.file_cache[cache_key] = {
                    "filename": filename_with_extension,
                    "content": data_url,
                    "is_image": True,
                    "original_filename": file.filename,
                    "mime_type": mime_type,
                    "size": len(file_content_bytes)
                }
            else:
                if not file_extension:
                    return jsonify({"error": f"Unsupported file type: {mime_type}"}), 400

                file_content = extract_text_from_file(file_content_bytes, mime_type)

                if file_content.startswith("Error:"):
                    return jsonify({"error": file_content}), 400

                current_app.file_cache[cache_key] = {
                    "filename": filename_with_extension,
                    "content": file_content,
                    "is_image": False,
                    "original_filename": file.filename,
                    "mime_type": mime_type,
                    "size": len(file_content_bytes)
                }

            return jsonify({
                "message": "File staged successfully",
                "filename": filename_with_extension,
                "size": len(file_content_bytes),
                "type": mime_type
            }), 200

    except Exception as e:
        logging.error(f"File upload error: {e}", exc_info=True)
        return jsonify({"error": f"File upload failed: {str(e)}"}), 500

    return jsonify({"error": "File upload failed"}), 500

@file_bp.route('/upload/status', methods=['GET'])
@token_required
def upload_status(current_user):
    """Check if there's a staged file for the session."""
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({"error": "session_id is required"}), 400

    user_id = current_user['id']
    cache_key = f"{user_id}-{session_id}"

    if hasattr(current_app, 'file_cache') and cache_key in current_app.file_cache:
        file_data = current_app.file_cache[cache_key]
        return jsonify({
            "has_file": True,
            "filename": file_data.get("filename"),
            "size": file_data.get("size"),
            "type": file_data.get("mime_type")
        }), 200

    return jsonify({"has_file": False}), 200

@file_bp.route('/upload/clear', methods=['POST'])
@token_required
def clear_upload(current_user):
    """Clear staged file for the session."""
    data = request.json or {}
    session_id = data.get('session_id')
    if not session_id:
        return jsonify({"error": "session_id is required"}), 400

    user_id = current_user['id']
    cache_key = f"{user_id}-{session_id}"

    if hasattr(current_app, 'file_cache') and cache_key in current_app.file_cache:
        del current_app.file_cache[cache_key]
        return jsonify({"message": "File cleared successfully"}), 200

    return jsonify({"message": "No file to clear"}), 200
